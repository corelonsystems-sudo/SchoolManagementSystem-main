from django.contrib import admin, messages
from django.utils.html import format_html
from django.urls import reverse, path
from django.db.models import Count
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect
import json
from .models import Student, Document, Application, ApplicationDocument
from finance.models import AdmissionFee
from academics.models import StudyYear, semester


class ApplicationDocumentInline(admin.TabularInline):
    model = ApplicationDocument
    extra = 1
    fields = ['name', 'file']
    max_num = 10


@admin.register(Application)
class ApplicationAdmin(admin.ModelAdmin):
    list_display = ['first_name', 'last_name', 'course', 'intake', 'status', 'application_date', 'admit_action']
    list_filter = ['status', 'course', 'intake']
    search_fields = ['first_name', 'last_name', 'email', 'phone']
    date_hierarchy = 'application_date'
    readonly_fields = ['application_date']
    change_list_template = 'admin/admissions/application/change_list.html'

    fieldsets = [
        ('Application Status', {
            'fields': ['status', 'review_notes']
        }),
        ('Student Bio Data', {
            'fields': [
                'passport_photo', 'first_name', 'middle_name',
                'last_name', 'date_of_birth', 'gender',
                'religion', 'nationality'
            ]
        }),
        ('Contact and Address Details', {
            'fields': [
                'birth_district', 'subcounty', 'parish',
                'village', 'phone', 'email'
            ]
        }),
        ('Parents Information', {
            'fields': [
                'father_name', 'father_phone',
                'mother_name', 'mother_phone'
            ]
        }),
        ('Sponsor Information', {
            'fields': ['sponsor_name']
        }),
        ('Previous School Information', {
            'fields': ['former_school', 'former_school_district']
        }),
        ('Academic Information', {
            'fields': ['course', 'intake', 'year_of_admission', 'NSIN']
        }),
    ]
    inlines = [ApplicationDocumentInline]

    def admit_action(self, obj):
        buttons = []
        try:
            has_student = obj.admitted_student.exists()
        except:
            has_student = False
        
        if obj.status == 'PENDING' or obj.status == 'UNDER_REVIEW':
            buttons.append(format_html(
                '<a href="{}" class="button" style="background-color:#28a745;margin-right:5px;">Admit</a>',
                reverse('admin:admit_application', args=[obj.id])
            ))
            buttons.append(format_html(
                '<a href="{}" class="button" style="background-color:#dc3545;">Reject</a>',
                reverse('admin:reject_application', args=[obj.id])
            ))
        elif obj.status == 'ACCEPTED' and not has_student:
            buttons.append(format_html(
                '<a href="{}" class="button" style="background-color:#28a745;">Admit</a>',
                reverse('admin:admit_application', args=[obj.id])
            ))
        elif obj.status == 'ADMITTED' or has_student:
            return format_html('<span style="color:green;font-weight:bold;">Admitted</span>')
        elif obj.status == 'REJECTED':
            return format_html('<span style="color:red;font-weight:bold;">Rejected</span>')
        return format_html(' '.join(buttons)) if buttons else '-'
    admit_action.short_description = 'Actions'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:application_id>/admit/', self.admin_site.admin_view(self.admit_application_view), name='admit_application'),
            path('<int:application_id>/reject/', self.admin_site.admin_view(self.reject_application_view), name='reject_application'),
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), name='import_applications_excel'),
            path('download-template/', self.admin_site.admin_view(self.download_template_view), name='download_application_template'),
            path('bulk-add-old/', self.admin_site.admin_view(self.bulk_add_old_view), name='bulk_add_old_applications'),
        ]
        return custom_urls + urls

    def import_excel_view(self, request):
        import openpyxl
        from io import BytesIO
        from academics.models import Course, Intake
        from datetime import datetime
        from django.http import JsonResponse

        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                return JsonResponse({'success': False, 'error': 'Please select an Excel file to upload.'})

            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({'success': False, 'error': 'Only .xlsx or .xls files are supported.'})

            try:
                file_content = excel_file.read()
                wb = openpyxl.load_workbook(BytesIO(file_content))
                ws = wb.active
            except Exception as e:
                return JsonResponse({'success': False, 'error': f'Error reading Excel file: {str(e)}'})

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return JsonResponse({'success': False, 'error': 'The Excel file has no data rows (only headers or empty).'})

            header_row = [str(cell).strip().lower().replace(' ', '_') if cell else '' for cell in rows[0]]

            # Map Excel column headers to model field names
            field_map = {
                'first_name': 'first_name',
                'middle_name': 'middle_name',
                'last_name': 'last_name',
                'date_of_birth': 'date_of_birth',
                'dob': 'date_of_birth',
                'gender': 'gender',
                'religion': 'religion',
                'nationality': 'nationality',
                'birth_district': 'birth_district',
                'subcounty': 'subcounty',
                'parish': 'parish',
                'village': 'village',
                'phone': 'phone',
                'email': 'email',
                'father_name': 'father_name',
                'father_phone': 'father_phone',
                'mother_name': 'mother_name',
                'mother_phone': 'mother_phone',
                'sponsor_name': 'sponsor_name',
                'former_school': 'former_school',
                'former_school_district': 'former_school_district',
                'course': 'course',
                'course_short_name': 'course_short_name',
                'intake': 'intake',
                'intake_name': 'intake_name',
                'year_of_admission': 'year_of_admission',
                'nsin': 'NSIN',
            }

            col_indices = {}
            for idx, header in enumerate(header_row):
                if header in field_map:
                    col_indices[field_map[header]] = idx

            required_fields = ['first_name', 'last_name', 'date_of_birth', 'gender',
                               'religion', 'nationality', 'birth_district', 'subcounty',
                               'parish', 'village', 'phone', 'email',
                               'father_name', 'father_phone', 'mother_name', 'mother_phone',
                               'sponsor_name', 'year_of_admission']
            missing = [f for f in required_fields if f not in col_indices]
            if missing:
                return JsonResponse({'success': False, 'error': f'Missing required columns: {", ".join(missing)}. '
                                    f'Expected columns: first_name, last_name, date_of_birth, gender, religion, '
                                    f'nationality, birth_district, subcounty, parish, village, phone, email, '
                                    f'father_name, father_phone, mother_name, mother_phone, sponsor_name, '
                                    f'course (or course_short_name), intake (or intake_name), year_of_admission.'})

            # Resolve course and intake lookup columns
            course_col = col_indices.get('course')
            course_short_col = col_indices.get('course_short_name')
            intake_col = col_indices.get('intake')
            intake_name_col = col_indices.get('intake_name')

            if not course_col and not course_short_col:
                return JsonResponse({'success': False, 'error': 'Missing course column. Provide either "course" (full name) or "course_short_name" (e.g., CS).'})

            if not intake_col and not intake_name_col:
                return JsonResponse({'success': False, 'error': 'Missing intake column. Provide either "intake" (ID) or "intake_name" (e.g., January 2026 Intake).'})

            created_count = 0
            error_rows = []

            for row_num, row in enumerate(rows[1:], start=2):
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue

                try:
                    def get_val(field_name):
                        idx = col_indices.get(field_name)
                        if idx is None or idx >= len(row):
                            return ''
                        val = row[idx]
                        if val is None:
                            return ''
                        return str(val).strip()

                    # Resolve course
                    course = None
                    if course_short_col is not None:
                        short_val = row[course_short_col]
                        if short_val:
                            course = Course.objects.filter(short_name__iexact=str(short_val).strip()).first()
                    if not course and course_col is not None:
                        course_val = row[course_col]
                        if course_val:
                            course = Course.objects.filter(name__iexact=str(course_val).strip()).first()
                    if not course:
                        error_rows.append(f'Row {row_num}: Course not found.')
                        continue

                    # Resolve intake
                    intake = None
                    if intake_name_col is not None:
                        intake_val = row[intake_name_col]
                        if intake_val:
                            intake = Intake.objects.filter(name__iexact=str(intake_val).strip()).first()
                    if not intake and intake_col is not None:
                        intake_val = row[intake_col]
                        if intake_val:
                            try:
                                intake = Intake.objects.get(id=int(intake_val))
                            except (ValueError, Intake.DoesNotExist):
                                pass
                    if not intake:
                        error_rows.append(f'Row {row_num}: Intake not found.')
                        continue

                    # Parse date_of_birth - handle multiple formats
                    dob_raw = get_val('date_of_birth')
                    dob = None
                    if isinstance(row[col_indices['date_of_birth']], datetime):
                        dob = row[col_indices['date_of_birth']].date()
                    elif dob_raw:
                        # Try multiple date formats
                        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']
                        for fmt in date_formats:
                            try:
                                dob = datetime.strptime(dob_raw, fmt).date()
                                break
                            except ValueError:
                                continue
                        if not dob:
                            error_rows.append(f'Row {row_num}: Invalid date format "{dob_raw}". Use YYYY-MM-DD.')
                            continue

                    # Validate gender
                    gender_val = get_val('gender').upper()
                    if gender_val not in ('M', 'F'):
                        if gender_val.startswith('M'):
                            gender_val = 'M'
                        elif gender_val.startswith('F'):
                            gender_val = 'F'
                        else:
                            error_rows.append(f'Row {row_num}: Invalid gender "{get_val("gender")}". Use M or F.')
                            continue

                    nsin_val = get_val('NSIN')

                    Application.objects.create(
                        first_name=get_val('first_name'),
                        middle_name=get_val('middle_name'),
                        last_name=get_val('last_name'),
                        date_of_birth=dob,
                        gender=gender_val,
                        religion=get_val('religion'),
                        nationality=get_val('nationality'),
                        birth_district=get_val('birth_district'),
                        subcounty=get_val('subcounty'),
                        parish=get_val('parish'),
                        village=get_val('village'),
                        phone=get_val('phone'),
                        email=get_val('email'),
                        father_name=get_val('father_name'),
                        father_phone=get_val('father_phone'),
                        mother_name=get_val('mother_name'),
                        mother_phone=get_val('mother_phone'),
                        sponsor_name=get_val('sponsor_name'),
                        former_school=get_val('former_school'),
                        former_school_district=get_val('former_school_district'),
                        course=course,
                        intake=intake,
                        year_of_admission=get_val('year_of_admission'),
                        NSIN=nsin_val,
                        status='PENDING',
                    )
                    created_count += 1
                except Exception as e:
                    error_rows.append(f'Row {row_num}: {str(e)}')

            result = {'success': True, 'created_count': created_count}
            if error_rows:
                result['errors'] = error_rows[:10]
                result['error_count'] = len(error_rows)
            return JsonResponse(result)

        # GET request - return partial template for modal
        context = {
            'opts': self.model._meta,
        }
        return render(request, 'admin/admissions/application/import_excel_partial.html', context)

    def download_template_view(self, request):
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
        from django.http import HttpResponse
        from academics.models import Course, Intake

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Application Template'

        headers = [
            'first_name', 'middle_name', 'last_name', 'date_of_birth', 'gender',
            'religion', 'nationality', 'birth_district', 'subcounty', 'parish',
            'village', 'phone', 'email', 'father_name', 'father_phone',
            'mother_name', 'mother_phone', 'sponsor_name', 'former_school',
            'former_school_district', 'course_short_name', 'course',
            'intake_name', 'intake', 'year_of_admission', 'NSIN'
        ]

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Add a sample row
        ws.cell(row=2, column=1, value='John')
        ws.cell(row=2, column=2, value='Doe')
        ws.cell(row=2, column=3, value='Smith')
        ws.cell(row=2, column=4, value='2005-01-15')
        ws.cell(row=2, column=5, value='M')
        ws.cell(row=2, column=6, value='Christian')
        ws.cell(row=2, column=7, value='Ugandan')
        ws.cell(row=2, column=8, value='Kampala')
        ws.cell(row=2, column=9, value='Central')
        ws.cell(row=2, column=10, value='Makerere')
        ws.cell(row=2, column=11, value='Kikoni')
        ws.cell(row=2, column=12, value='256700123456')
        ws.cell(row=2, column=13, value='john.doe@example.com')
        ws.cell(row=2, column=14, value='James Doe')
        ws.cell(row=2, column=15, value='256700123457')
        ws.cell(row=2, column=16, value='Mary Doe')
        ws.cell(row=2, column=17, value='256700123458')
        ws.cell(row=2, column=18, value='Mr. Sponsor')
        ws.cell(row=2, column=19, value='St. Marys Secondary')
        ws.cell(row=2, column=20, value='Kampala')
        ws.cell(row=2, column=21, value='CS')
        ws.cell(row=2, column=22, value='')
        ws.cell(row=2, column=23, value='January 2026 Intake')
        ws.cell(row=2, column=24, value='')
        ws.cell(row=2, column=25, value='2026')
        ws.cell(row=2, column=26, value='')

        # Add data validation for course_short_name (column 21)
        courses = list(Course.objects.values_list('short_name', flat=True).order_by('short_name'))
        if courses:
            course_dv = DataValidation(type='list', formula1=f'"{",".join(courses)}"', allow_blank=True)
            course_dv.error = 'Please select a valid course short name from the dropdown'
            course_dv.errorTitle = 'Invalid Course'
            course_dv.prompt = 'Select a course from the dropdown'
            course_dv.promptTitle = 'Course Selection'
            ws.add_data_validation(course_dv)
            course_dv.add('U2:U1000')  # Apply to course_short_name column

        # Add data validation for intake_name (column 23)
        intakes = list(Intake.objects.values_list('name', flat=True).order_by('-start_date'))
        if intakes:
            intake_dv = DataValidation(type='list', formula1=f'"{",".join(intakes)}"', allow_blank=True)
            intake_dv.error = 'Please select a valid intake from the dropdown'
            intake_dv.errorTitle = 'Invalid Intake'
            intake_dv.prompt = 'Select an intake from the dropdown'
            intake_dv.promptTitle = 'Intake Selection'
            ws.add_data_validation(intake_dv)
            intake_dv.add('W2:W1000')  # Apply to intake_name column

        # Add data validation for gender (column 5)
        gender_dv = DataValidation(type='list', formula1='"M,F"', allow_blank=False)
        gender_dv.error = 'Please select M or F'
        gender_dv.errorTitle = 'Invalid Gender'
        gender_dv.prompt = 'Select M (Male) or F (Female)'
        gender_dv.promptTitle = 'Gender Selection'
        ws.add_data_validation(gender_dv)
        gender_dv.add('E2:E1000')  # Apply to gender column

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="application_import_template.xlsx"'
        wb.save(response)
        return response

    def bulk_add_old_view(self, request):
        from academics.models import Course, Intake
        from datetime import datetime
        from django.http import JsonResponse

        courses = Course.objects.all().order_by('name')
        intakes = Intake.objects.all().select_related('academic_year').order_by('-start_date')
        status_choices = Application.STATUS_CHOICES

        if request.method == 'POST':
            row_count = int(request.POST.get('row_count', 0))
            created_count = 0
            errors = []

            for i in range(row_count):
                first_name = request.POST.get(f'first_name_{i}', '').strip()
                last_name = request.POST.get(f'last_name_{i}', '').strip()

                if not first_name and not last_name:
                    continue

                try:
                    course_id = request.POST.get(f'course_{i}')
                    intake_id = request.POST.get(f'intake_{i}')

                    if not course_id or not intake_id:
                        errors.append(f'Row {i+1}: Course and Intake are required.')
                        continue

                    dob_str = request.POST.get(f'date_of_birth_{i}', '').strip()
                    dob = None
                    if dob_str:
                        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']
                        for fmt in date_formats:
                            try:
                                dob = datetime.strptime(dob_str, fmt).date()
                                break
                            except ValueError:
                                continue

                    gender = request.POST.get(f'gender_{i}', '').strip().upper()
                    if gender and gender not in ('M', 'F'):
                        if gender.startswith('M'):
                            gender = 'M'
                        elif gender.startswith('F'):
                            gender = 'F'

                    status = request.POST.get(f'status_{i}', 'PENDING')
                    if status not in dict(status_choices):
                        status = 'PENDING'

                    Application.objects.create(
                        first_name=first_name,
                        middle_name=request.POST.get(f'middle_name_{i}', '').strip(),
                        last_name=last_name,
                        date_of_birth=dob,
                        gender=gender or 'M',
                        religion=request.POST.get(f'religion_{i}', '').strip(),
                        nationality=request.POST.get(f'nationality_{i}', '').strip(),
                        birth_district=request.POST.get(f'birth_district_{i}', '').strip(),
                        subcounty=request.POST.get(f'subcounty_{i}', '').strip(),
                        parish=request.POST.get(f'parish_{i}', '').strip(),
                        village=request.POST.get(f'village_{i}', '').strip(),
                        phone=request.POST.get(f'phone_{i}', '').strip(),
                        email=request.POST.get(f'email_{i}', '').strip(),
                        father_name=request.POST.get(f'father_name_{i}', '').strip(),
                        father_phone=request.POST.get(f'father_phone_{i}', '').strip(),
                        mother_name=request.POST.get(f'mother_name_{i}', '').strip(),
                        mother_phone=request.POST.get(f'mother_phone_{i}', '').strip(),
                        sponsor_name=request.POST.get(f'sponsor_name_{i}', '').strip(),
                        former_school=request.POST.get(f'former_school_{i}', '').strip(),
                        former_school_district=request.POST.get(f'former_school_district_{i}', '').strip(),
                        course_id=course_id,
                        intake_id=intake_id,
                        year_of_admission=request.POST.get(f'year_of_admission_{i}', '').strip(),
                        NSIN=request.POST.get(f'NSIN_{i}', '').strip(),
                        status=status,
                    )
                    created_count += 1
                except Exception as e:
                    errors.append(f'Row {i+1}: {str(e)}')

            result = {'success': True, 'created_count': created_count}
            if errors:
                result['errors'] = errors[:10]
                result['error_count'] = len(errors)
            return JsonResponse(result)

        context = {
            'opts': self.model._meta,
            'courses': courses,
            'intakes': intakes,
            'status_choices': status_choices,
        }
        return render(request, 'admin/admissions/application/bulk_add_old_partial.html', context)

    def admit_application_view(self, request, application_id):
        from django.shortcuts import get_object_or_404
        application = get_object_or_404(Application, id=application_id)

        if application.status == 'REJECTED':
            messages.error(request, 'Rejected applications cannot be admitted.')
            return redirect('admin:admissions_application_changelist')

        try:
            has_student = application.admitted_student.exists()
        except:
            has_student = False

        if has_student:
            messages.warning(request, 'This application has already been admitted.')
            return redirect('admin:admissions_application_changelist')

        # Create a Student from the application
        student = Student.objects.create(
            application=application,
            first_name=application.first_name,
            middle_name=application.middle_name,
            last_name=application.last_name,
            date_of_birth=application.date_of_birth,
            gender=application.gender,
            religion=application.religion,
            nationality=application.nationality,
            passport_photo=application.passport_photo,
            birth_district=application.birth_district,
            subcounty=application.subcounty,
            parish=application.parish,
            village=application.village,
            phone=application.phone,
            email=application.email,
            father_name=application.father_name,
            father_phone=application.father_phone,
            mother_name=application.mother_name,
            mother_phone=application.mother_phone,
            sponsor_name=application.sponsor_name,
            former_school=application.former_school,
            former_school_district=application.former_school_district,
            course=application.course,
            year_of_admission=application.year_of_admission,
            NSIN=application.NSIN,
            enrollment_status='ENROLLED'
        )

        # Update application status
        application.status = 'ADMITTED'
        application.save()

        # Copy application documents to student documents
        for app_doc in application.documents.all():
            Document.objects.create(
                student=student,
                name=app_doc.name,
                file=app_doc.file
            )

        messages.success(request, f'Student {student.first_name} {student.last_name} has been admitted successfully.')
        return redirect('admin:admissions_student_change', student.id)

    def reject_application_view(self, request, application_id):
        from django.shortcuts import get_object_or_404
        application = get_object_or_404(Application, id=application_id)

        try:
            has_student = application.admitted_student.exists()
        except:
            has_student = False

        if application.status == 'ADMITTED' or has_student:
            messages.error(request, 'Admitted applications cannot be rejected.')
            return redirect('admin:admissions_application_changelist')

        if application.status == 'REJECTED':
            messages.warning(request, 'This application has already been rejected.')
            return redirect('admin:admissions_application_changelist')

        # Update application status
        application.status = 'REJECTED'
        application.save()

        messages.success(request, f'Application from {application.first_name} {application.last_name} has been rejected.')
        return redirect('admin:admissions_application_changelist')


class DocumentInline(admin.TabularInline):
    model = Document
    extra = 1
    fields = ['name', 'file']
    max_num = 10

class AdmissionFeeInline(admin.TabularInline):
    model = AdmissionFee
    extra = 0
    fields = ['payment_type_display', 'frequency_display', 'default_amount_display', 'custom_amount', 'is_active']
    readonly_fields = ['payment_type_display', 'frequency_display', 'default_amount_display']
    autocomplete_fields = []

    def payment_type_display(self, obj):
        return obj.payment_type.name
    payment_type_display.short_description = 'Payment Type'

    def frequency_display(self, obj):
        return obj.payment_type.get_frequency_display()
    frequency_display.short_description = 'Frequency'

    def default_amount_display(self, obj):
        return obj.payment_type.default_amount
    default_amount_display.short_description = 'Default Amount'

from django.utils.html import format_html

@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    actions = []
    change_list_template = 'admin/admissions/student/change_list.html'
    list_display = ['first_name', 'last_name', 'admission_number', 'course', 'get_intake', 'year_of_admission', 'student_actions']
    search_fields = ['first_name', 'last_name', 'admission_number']
    list_filter = ['course', 'year_of_admission', 'enrollment_status']
    readonly_fields = ['admission_number', 'preview_passport_photo']
    fieldsets = [
        ('Student Bio Data', {
            'fields': [
                'preview_passport_photo', 'passport_photo', 'first_name', 'middle_name',
                'last_name', 'date_of_birth', 'gender',
                'religion', 'nationality'
            ]
        }),
        ('Contact and Address Details', {
            'fields': [
                'birth_district', 'subcounty', 'parish',
                'village', 'phone', 'email'
            ]
        }),
        ('Parents Information', {
            'fields': [
                'father_name', 'father_phone',
                'mother_name', 'mother_phone'
            ]
        }),
        ('Sponsor Information', {
            'fields': ['sponsor_name']
        }),
        ('Previous School Information', {
            'fields': ['former_school', 'former_school_district']
        }),
        ('Academic Information', {
            'fields': ['course', 'application', 'admission_number', 'year_of_admission', 'reporting', 'NSIN', 'enrollment_status']
        }),
    ]
    
    inlines = [DocumentInline, AdmissionFeeInline]
    search_fields = ['first_name', 'last_name', 'admission_number']
    list_filter = ['course', 'year_of_admission', 'enrollment_status']
    readonly_fields = ['admission_number', 'preview_passport_photo']

    def student_actions(self, obj):
        buttons = []
        # Enroll button - opens popup
        buttons.append(format_html(
            '<a href="#" onclick="openEnrollModal(event, \'{}\')" class="button" style="background-color:#007bff;margin-right:5px;">Enroll</a>',
            reverse('admin:enroll_student', args=[obj.id])
        ))
        # Print button
        buttons.append(format_html(
            '<a href="{}" target="_blank" class="button" style="background-color:#28a745;margin-right:5px;">Print</a>',
            reverse('student_admission_form', args=[obj.id])
        ))
        # Download button
        buttons.append(format_html(
            '<a href="{}" target="_blank" class="button" style="background-color:#17a2b8;margin-right:5px;">Download</a>',
            reverse('admission_letter', args=[obj.id])
        ))
        # View button
        buttons.append(format_html(
            '<a href="{}" class="button" style="background-color:#6c757d;">View</a>',
            reverse('admin:admissions_student_change', args=[obj.id])
        ))
        return format_html(' '.join(buttons))
    student_actions.short_description = 'Actions'

    def get_intake(self, obj):
        if obj.application and obj.application.intake:
            return obj.application.intake
        return '-'
    get_intake.short_description = 'Intake'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:student_id>/enroll/', self.admin_site.admin_view(self.enroll_student_view), name='enroll_student'),
            path('<int:student_id>/enroll/submit/', self.admin_site.admin_view(self.enroll_student_submit), name='enroll_student_submit'),
        ]
        return custom_urls + urls

    def enroll_student_view(self, request, student_id):
        from django.shortcuts import get_object_or_404
        student = get_object_or_404(Student, id=student_id)
        
        context = {
            **self.admin_site.each_context(request),
            'student': student,
            'opts': self.model._meta,
            'title': f'Enroll {student.first_name} {student.last_name}',
            'study_years': StudyYear.objects.all(),
            'semesters': semester.objects.all(),
        }
        return render(request, 'admin/admissions/student/enroll.html', context)

    def enroll_student_submit(self, request, student_id):
        from django.shortcuts import get_object_or_404
        from academics.models import Enrollment, Student as AcademicsStudent
        admissions_student = get_object_or_404(Student, id=student_id)
        
        study_year_id = request.POST.get('study_year')
        semester_id = request.POST.get('semester')
        
        if not study_year_id or not semester_id:
            messages.error(request, 'Please select both study year and semester.')
            return redirect('admin:enroll_student', student_id)
        
        try:
            study_year = StudyYear.objects.get(id=study_year_id)
            sem = semester.objects.get(id=semester_id)
        except (StudyYear.DoesNotExist, semester.DoesNotExist):
            messages.error(request, 'Invalid study year or semester selected.')
            return redirect('admin:enroll_student', student_id)
        
        # Check if academics student already exists
        try:
            academics_student = AcademicsStudent.objects.get(admission=admissions_student)
            # Update intake if not set
            if not academics_student.intake and admissions_student.application and admissions_student.application.intake:
                academics_student.intake = admissions_student.application.intake
                academics_student.save()
        except AcademicsStudent.DoesNotExist:
            # Create academics student record
            academics_student = AcademicsStudent.objects.create(
                admission=admissions_student,
                reg_number=admissions_student.admission_number,
                intake=admissions_student.application.intake if admissions_student.application and admissions_student.application.intake else None
            )
        
        # Create enrollment record
        Enrollment.objects.create(
            student=academics_student,
            study_year=study_year,
            semester=sem
        )
        
        # Update student enrollment status
        admissions_student.enrollment_status = 'ENROLLED'
        admissions_student.save()
        
        context = {
            'success': True,
            'message': f'{admissions_student.first_name} {admissions_student.last_name} has been enrolled successfully.'
        }
        return render(request, 'admin/admissions/student/enroll_success.html', context)

    def preview_passport_photo(self, obj):
        if obj.passport_photo:
            return format_html('<img src="{}" style="width: 100px; height: auto;" />', obj.passport_photo.url)
        return "No photo uploaded"

    preview_passport_photo.short_description = "Passport Photo Preview"

    def print_admission_letter(self, obj):
        url = reverse('admission_letter', args=[obj.id])
        return format_html('<a href="{}" target="_blank">Download Admission Letter</a>', url)

    print_admission_letter.short_description = 'Download Admission Letter'

    def enrollment_status_badge(self, obj):
        colors = {
            'ENROLLED': '#28a745',
            'PENDING': '#ffc107',
            'DEFERRED': '#6c757d',
            'COMPLETED': '#007bff',
            'WITHDRAWN': '#dc3545',
        }
        color = colors.get(obj.enrollment_status, '#6c757d')
        return format_html(
            '<span style="background-color:{}; color:#fff; padding:2px 8px; border-radius:3px; font-size:0.8em;">{}</span>',
            color, obj.get_enrollment_status_display()
        )

    enrollment_status_badge.short_description = 'Enrollment Status'
    enrollment_status_badge.admin_order_field = 'enrollment_status'

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        study_years = StudyYear.objects.all()
        semesters = semester.objects.all()
        extra_context['study_years'] = study_years
        extra_context['semesters'] = semesters
        response = super().changelist_view(request, extra_context=extra_context)
        return response

@admin.register(Document)
class DocumentAdmin(admin.ModelAdmin):
    list_display = ['name', 'student', 'uploaded_at']
    search_fields = ['name', 'student__first_name', 'student__last_name']

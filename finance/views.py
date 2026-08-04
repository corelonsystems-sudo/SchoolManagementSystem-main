from django.shortcuts import render
from django.db.models import Sum, Q
from .models import Payment, Ledger, PaymentType
from academics.models import StudyYear, semester, Course, Intake, Student as AcademicsStudent
from admissions.models import Student as AdmissionsStudent
from datetime import datetime
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl

def payment_list(request):
    payments = Payment.objects.all()
    return render(request, 'finance/payment_list.html', {'payments': payments})


def financial_reports(request):
    # Get filter values from request
    report_type = request.GET.get('report_type', 'payments')
    study_year_id = request.GET.get('study_year')
    semester_id = request.GET.get('semester')
    intake_id = request.GET.get('intake')
    course_id = request.GET.get('course')
    payment_type_id = request.GET.get('payment_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Initialize data containers
    payments = []
    ledgers = []
    balances = []
    debts = []
    
    # Common filter conditions
    study_year_filter = {'study_year_id': study_year_id} if study_year_id else {}
    semester_filter = {'semester_id': semester_id} if semester_id else {}
    intake_filter = {'student__intake_id': intake_id} if intake_id else {}
    course_filter = {'student__admission__course_id': course_id} if course_id else {}
    
    # Handle different report types
    if report_type == 'payments':
        payments = Payment.objects.select_related(
            'enrollment__student__admission',
            'enrollment__student__intake',
            'enrollment__study_year',
            'enrollment__semester',
            'payment_type',
            'ledger'
        ).all()
        
        if study_year_id:
            payments = payments.filter(enrollment__study_year_id=study_year_id)
        if semester_id:
            payments = payments.filter(enrollment__semester_id=semester_id)
        if intake_id:
            payments = payments.filter(enrollment__student__intake_id=intake_id)
        if course_id:
            payments = payments.filter(enrollment__student__admission__course_id=course_id)
        if payment_type_id:
            payments = payments.filter(payment_type_id=payment_type_id)
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d')
                payments = payments.filter(date__gte=date_from)
            except ValueError:
                pass
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d')
                payments = payments.filter(date__lte=date_to)
            except ValueError:
                pass
        
        total_amount = payments.aggregate(total=Sum('amount'))['total'] or 0
        total_count = payments.count()
        
    elif report_type == 'ledgers':
        ledgers = Ledger.objects.select_related('student__admission', 'study_year', 'semester').all()
        
        if study_year_id:
            ledgers = ledgers.filter(study_year_id=study_year_id)
        if semester_id:
            ledgers = ledgers.filter(semester_id=semester_id)
        if intake_id:
            ledgers = ledgers.filter(student__intake_id=intake_id)
        if course_id:
            ledgers = ledgers.filter(student__admission__course_id=course_id)
        
        total_required = ledgers.aggregate(total=Sum('required_amount'))['total'] or 0
        ledger_count = ledgers.count()
        
    elif report_type == 'balances':
        ledgers = Ledger.objects.select_related('student__admission', 'study_year', 'semester').all()
        
        if study_year_id:
            ledgers = ledgers.filter(study_year_id=study_year_id)
        if semester_id:
            ledgers = ledgers.filter(semester_id=semester_id)
        if intake_id:
            ledgers = ledgers.filter(student__intake_id=intake_id)
        if course_id:
            ledgers = ledgers.filter(student__admission__course_id=course_id)
        
        balances = []
        for ledger in ledgers:
            paid = Payment.objects.filter(
                enrollment__student=ledger.student,
                enrollment__study_year=ledger.study_year,
                enrollment__semester=ledger.semester
            ).aggregate(total=Sum('amount'))['total'] or 0
            balance = ledger.required_amount - paid
            balances.append({
                'ledger': ledger,
                'paid': paid,
                'balance': balance
            })
        
        ledger_count = len(balances)
        total_required = sum(b['ledger'].required_amount for b in balances)
        total_paid = sum(b['paid'] for b in balances)
        total_balance = total_required - total_paid
        
    elif report_type == 'debts':
        ledgers = Ledger.objects.select_related('student__admission', 'study_year', 'semester').all()
        
        if study_year_id:
            ledgers = ledgers.filter(study_year_id=study_year_id)
        if semester_id:
            ledgers = ledgers.filter(semester_id=semester_id)
        if intake_id:
            ledgers = ledgers.filter(student__intake_id=intake_id)
        if course_id:
            ledgers = ledgers.filter(student__admission__course_id=course_id)
        
        debts = []
        for ledger in ledgers:
            paid = Payment.objects.filter(
                enrollment__student=ledger.student,
                enrollment__study_year=ledger.study_year,
                enrollment__semester=ledger.semester
            ).aggregate(total=Sum('amount'))['total'] or 0
            balance = ledger.required_amount - paid
            if balance > 0:
                debts.append({
                    'ledger': ledger,
                    'paid': paid,
                    'balance': balance
                })
        
        ledger_count = len(debts)
        total_required = sum(d['ledger'].required_amount for d in debts)
        total_paid = sum(d['paid'] for d in debts)
        total_balance = total_required - total_paid

    # Get filter options
    study_years = StudyYear.objects.all().order_by('-year')
    semesters = semester.objects.all().order_by('semester')
    intakes = Intake.objects.all().order_by('-start_date')
    courses = Course.objects.all().order_by('name')
    payment_types = PaymentType.objects.all().order_by('name')

    # Build filter dict for template
    filters = {
        'report_type': report_type,
        'study_year': study_year_id,
        'semester': semester_id,
        'intake': intake_id,
        'course': course_id,
        'payment_type': payment_type_id,
        'date_from': date_from,
        'date_to': date_to,
    }

    # Calculate totals for summary cards
    if report_type == 'payments':
        total_required = 0
        total_paid_ledger = 0
        total_balance = 0
        ledger_count = 0
    elif report_type == 'ledgers':
        total_amount = 0
        total_count = 0
        total_paid_ledger = 0
        total_balance = 0
    elif report_type == 'balances':
        total_amount = 0
        total_count = 0
        total_paid_ledger = total_paid
        total_balance = total_balance
    elif report_type == 'debts':
        total_amount = 0
        total_count = 0
        total_paid_ledger = total_paid
        total_balance = total_balance

    return render(request, 'finance/payment_reports.html', {
        'report_type': report_type,
        'payments': payments,
        'ledgers': ledgers,
        'balances': balances,
        'debts': debts,
        'total_amount': total_amount if report_type == 'payments' else 0,
        'total_count': total_count if report_type == 'payments' else 0,
        'total_required': total_required if report_type != 'payments' else 0,
        'total_paid_ledger': total_paid_ledger if report_type != 'payments' else 0,
        'total_balance': total_balance if report_type != 'payments' else 0,
        'ledger_count': ledger_count if report_type != 'payments' else 0,
        'study_years': study_years,
        'semesters': semesters,
        'intakes': intakes,
        'courses': courses,
        'payment_types': payment_types,
        'filters': filters,
    })


def download_financial_report(request):
    # Get filter values from request
    study_year_id = request.GET.get('study_year')
    semester_id = request.GET.get('semester')
    intake_id = request.GET.get('intake')
    course_id = request.GET.get('course')
    payment_type_id = request.GET.get('payment_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Build queryset
    payments = Payment.objects.select_related(
        'enrollment__student__admission',
        'enrollment__student__intake',
        'enrollment__study_year',
        'enrollment__semester',
        'payment_type',
        'ledger'
    ).all()

    # Apply filters
    if study_year_id:
        payments = payments.filter(enrollment__study_year_id=study_year_id)
    if semester_id:
        payments = payments.filter(enrollment__semester_id=semester_id)
    if intake_id:
        payments = payments.filter(enrollment__student__intake_id=intake_id)
    if course_id:
        payments = payments.filter(enrollment__student__admission__course_id=course_id)
    if payment_type_id:
        payments = payments.filter(payment_type_id=payment_type_id)
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            payments = payments.filter(date__gte=date_from)
        except ValueError:
            pass
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            payments = payments.filter(date__lte=date_to)
        except ValueError:
            pass

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payment Report'

    # Headers
    headers = ['Date', 'Student Name', 'Student ID', 'Course', 'Intake', 'Academic Year', 'Semester', 'Payment Type', 'Amount', 'Ledger Number']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Data rows
    for row_num, payment in enumerate(payments, 2):
        ws.cell(row=row_num, column=1, value=payment.date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_num, column=2, value=f"{payment.enrollment.student.admission.first_name} {payment.enrollment.student.admission.last_name}")
        ws.cell(row=row_num, column=3, value=payment.enrollment.student.display_id)
        ws.cell(row=row_num, column=4, value=payment.enrollment.student.admission.course.name)
        ws.cell(row=row_num, column=5, value=payment.enrollment.student.intake.name if payment.enrollment.student.intake else '')
        ws.cell(row=row_num, column=6, value=payment.enrollment.study_year.year if payment.enrollment.study_year else '')
        ws.cell(row=row_num, column=7, value=payment.enrollment.semester.semester if payment.enrollment.semester else '')
        ws.cell(row=row_num, column=8, value=payment.payment_type.name)
        ws.cell(row=row_num, column=9, value=float(payment.amount))
        ws.cell(row=row_num, column=10, value=payment.ledger.ledger_number if payment.ledger else '')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payment_report.xlsx"'
    wb.save(response)
    return response


@csrf_exempt
def download_financial_summary(request):
    if request.method != 'POST':
        return HttpResponse('Method not allowed', status=405)
    
    # Get selected sheets
    selected_sheets = request.POST.getlist('sheets')
    
    # Get filter values
    study_year_id = request.POST.get('study_year')
    semester_id = request.POST.get('semester')
    intake_id = request.POST.get('intake')
    course_id = request.POST.get('course')
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Payments sheet
    if 'payments' in selected_sheets:
        ws = wb.create_sheet('Payments')
        headers = ['Date', 'Student Name', 'Student ID', 'Course', 'Intake', 'Academic Year', 'Semester', 'Payment Type', 'Amount', 'Ledger Number', 'Balance After Payment']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        payments = Payment.objects.select_related(
            'enrollment__student__admission',
            'enrollment__student__intake',
            'enrollment__study_year',
            'enrollment__semester',
            'payment_type',
            'ledger'
        ).all()
        
        if study_year_id:
            payments = payments.filter(enrollment__study_year_id=study_year_id)
        if semester_id:
            payments = payments.filter(enrollment__semester_id=semester_id)
        if intake_id:
            payments = payments.filter(enrollment__student__intake_id=intake_id)
        if course_id:
            payments = payments.filter(enrollment__student__admission__course_id=course_id)
        
        for row_num, payment in enumerate(payments, 2):
            balance_after = payment.ledger.balance if payment.ledger else 0
            ws.cell(row=row_num, column=1, value=payment.date.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=2, value=f"{payment.enrollment.student.admission.first_name} {payment.enrollment.student.admission.last_name}")
            ws.cell(row=row_num, column=3, value=payment.enrollment.student.display_id)
            ws.cell(row=row_num, column=4, value=payment.enrollment.student.admission.course.name)
            ws.cell(row=row_num, column=5, value=payment.enrollment.student.intake.name if payment.enrollment.student.intake else '')
            ws.cell(row=row_num, column=6, value=payment.enrollment.study_year.year if payment.enrollment.study_year else '')
            ws.cell(row=row_num, column=7, value=payment.enrollment.semester.semester if payment.enrollment.semester else '')
            ws.cell(row=row_num, column=8, value=payment.payment_type.name)
            ws.cell(row=row_num, column=9, value=float(payment.amount))
            ws.cell(row=row_num, column=10, value=payment.ledger.ledger_number if payment.ledger else '')
            ws.cell(row=row_num, column=11, value=float(balance_after))
    
    # Ledgers sheet
    if 'ledgers' in selected_sheets:
        ws = wb.create_sheet('Ledgers')
        headers = ['Ledger Number', 'Student Name', 'Student ID', 'Course', 'Intake', 'Academic Year', 'Semester', 'Required Amount', 'Total Paid', 'Balance', 'Generated On']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        ledgers = Ledger.objects.select_related('student__admission', 'study_year', 'semester').all()
        
        if study_year_id:
            ledgers = ledgers.filter(study_year_id=study_year_id)
        if semester_id:
            ledgers = ledgers.filter(semester_id=semester_id)
        if intake_id:
            ledgers = ledgers.filter(student__intake_id=intake_id)
        if course_id:
            ledgers = ledgers.filter(student__admission__course_id=course_id)
        
        for row_num, ledger in enumerate(ledgers, 2):
            ws.cell(row=row_num, column=1, value=ledger.ledger_number)
            ws.cell(row=row_num, column=2, value=f"{ledger.student.admission.first_name} {ledger.student.admission.last_name}")
            ws.cell(row=row_num, column=3, value=ledger.student.display_id)
            ws.cell(row=row_num, column=4, value=ledger.student.admission.course.name)
            ws.cell(row=row_num, column=5, value=ledger.student.intake.name if ledger.student.intake else '')
            ws.cell(row=row_num, column=6, value=ledger.study_year.year if ledger.study_year else '')
            ws.cell(row=row_num, column=7, value=ledger.semester.semester if ledger.semester else '')
            ws.cell(row=row_num, column=8, value=float(ledger.required_amount))
            ws.cell(row=row_num, column=9, value=float(ledger.total_paid))
            ws.cell(row=row_num, column=10, value=float(ledger.balance))
            ws.cell(row=row_num, column=11, value=ledger.generated_on.strftime('%Y-%m-%d %H:%M'))
    
    # Debts sheet
    if 'debts' in selected_sheets:
        ws = wb.create_sheet('Debts')
        headers = ['Student Name', 'Student ID', 'Course', 'Intake', 'Academic Year', 'Semester', 'Ledger Number', 'Required Amount', 'Total Paid', 'Outstanding Balance', 'Debt Status']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        ledgers = Ledger.objects.select_related('student__admission', 'study_year', 'semester').filter(balance__gt=0)
        
        if study_year_id:
            ledgers = ledgers.filter(study_year_id=study_year_id)
        if semester_id:
            ledgers = ledgers.filter(semester_id=semester_id)
        if intake_id:
            ledgers = ledgers.filter(student__intake_id=intake_id)
        if course_id:
            ledgers = ledgers.filter(student__admission__course_id=course_id)
        
        for row_num, ledger in enumerate(ledgers, 2):
            debt_status = 'Partial' if ledger.total_paid > 0 else 'Unpaid'
            ws.cell(row=row_num, column=1, value=f"{ledger.student.admission.first_name} {ledger.student.admission.last_name}")
            ws.cell(row=row_num, column=2, value=ledger.student.display_id)
            ws.cell(row=row_num, column=3, value=ledger.student.admission.course.name)
            ws.cell(row=row_num, column=4, value=ledger.student.intake.name if ledger.student.intake else '')
            ws.cell(row=row_num, column=5, value=ledger.study_year.year if ledger.study_year else '')
            ws.cell(row=row_num, column=6, value=ledger.semester.semester if ledger.semester else '')
            ws.cell(row=row_num, column=7, value=ledger.ledger_number)
            ws.cell(row=row_num, column=8, value=float(ledger.required_amount))
            ws.cell(row=row_num, column=9, value=float(ledger.total_paid))
            ws.cell(row=row_num, column=10, value=float(ledger.balance))
            ws.cell(row=row_num, column=11, value=debt_status)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="financial_summary.xlsx"'
    wb.save(response)
    return response
